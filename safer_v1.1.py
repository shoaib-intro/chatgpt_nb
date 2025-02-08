from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
import time
import openpyxl  # For Excel integration
from selenium.webdriver.common.action_chains import ActionChains
import re
import random
import yaml
import os
from dotenv import load_dotenv

# Load environment variables from .env.local file
load_dotenv('.env.local')

try:
    EMAIL_LOGIN_URL = os.getenv('EMAIL_LOGIN_URL', 'https://mail.hostinger.com/')
    EMAIL_USERNAME  = os.getenv('EMAIL_USERNAME', None)
    EMAIL_PASSWORD  = os.getenv('EMAIL_PASSWORD', None)
    ENABLE_EMAIL_SENDING = os.getenv('ENABLE_EMAIL_SENDING', False)

    MCMX_START = int(os.getenv('MCMX_START', None))
    MCMX_END = int(os.getenv('MCMX_END', None))

except Exception as e:
    print(f"Error: {e}")


def setup_browser():
    # Set up the Selenium WebDriver (Ensure you have the ChromeDriver installed and PATH set)
    chrome_options = Options()
    chrome_options.add_argument("--log-level=3")  # Suppress browser warnings
    chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])  # Suppress DevTools warnings

    service = Service()
    driver = webdriver.Chrome(service=service, options=chrome_options)
    driver.get("https://safer.fmcsa.dot.gov/CompanySnapshot.aspx")
    return driver


def search_company(driver, number):
    _wait = random.randint(5, 10)
    try:
        # Navigate back to the search page
        driver.get("https://safer.fmcsa.dot.gov/CompanySnapshot.aspx")

        # Select the "MC/MX Number" radio button
        try:
            mc_mx_radio = WebDriverWait(driver, _wait).until(
                EC.element_to_be_clickable((By.XPATH, "//input[@id='2' and @name='query_param' and @value='MC_MX']"))
            )
            driver.execute_script("arguments[0].scrollIntoView(true);", mc_mx_radio)  # Ensure element is visible
            mc_mx_radio.click()
        except Exception as e:
            print(f"MC SEARCHER: MC/MX Number radio button not found or not clickable: {e}")
            return False

        # Enter the number in the search box
        try:
            search_box = WebDriverWait(driver, _wait).until(
                EC.presence_of_element_located((By.XPATH, "//input[@id='4' and @name='query_string']"))
            )
            search_box.clear()
            search_box.send_keys(number)
        except Exception as e:
            print(f"MC SEARCHER: Search box not found or not interactable: {e}")
            return False

        # Click the "Search" button
        try:
            search_button = WebDriverWait(driver, _wait).until(
                EC.element_to_be_clickable((By.XPATH, "//input[@type='SUBMIT' and @value='Search']"))
            )
            search_button.click()
        except Exception as e:
            print(f"MC SEARCHER: Search button not found or not clickable: {e}")
            return False

        return True
    except Exception as e:
        print(f"MC SEARCHER: An unexpected error occurred during the search process: {e}")
        return False

def verify_details(driver):
    print("\n========== CARRIER-VERIFICATION Section Start ==========")
    _wait = random.randint(2, 5)
    
    # Check if "Record Not Found" is present on the page
    try:
        record_not_found = WebDriverWait(driver, _wait).until(
            EC.presence_of_element_located((By.XPATH, "//i[text()='Record Not Found']"))
        )
        print("CARRIER-VERIFICATION: Record Not Found. Skipping further processing.")
        return False
    except Exception:
        pass

    # Check if "Record Inactive" is present on the page
    try:
        record_inactive = WebDriverWait(driver, _wait).until(
            EC.presence_of_element_located((By.XPATH, "//i[text()='Record Inactive']"))
        )
        print("CARRIER-VERIFICATION: Record Inactive. Skipping further processing.")
        return False
    except Exception:
        pass

    # Verify the details on the loaded page
    try:
        # Verify "Entity Type"
        entity_type = WebDriverWait(driver, _wait).until(
            EC.presence_of_element_located((By.XPATH, "//th[a[text()='Entity Type:']]/following-sibling::td"))
        ).text.strip()
        if entity_type != "CARRIER":
            print(f"CARRIER-VERIFICATION: Entity Type verification failed: {entity_type}")
            return False

        # Verify "USDOT Status"
        usdot_status = WebDriverWait(driver, _wait).until(
            EC.presence_of_element_located((By.XPATH, "//th[a[text()='USDOT Status:']]/following-sibling::td"))
        ).text.strip()
        if usdot_status != "ACTIVE":
            print(f"CARRIER-VERIFICATION: USDOT Status verification failed: {usdot_status}")
            return False

        # Verify "Operating Authority Status"
        operating_authority_status = WebDriverWait(driver, _wait).until(
            EC.presence_of_element_located((By.XPATH, "//th[a[text()='Operating Authority Status:']]/following-sibling::td"))
        ).text.strip()
        if "AUTHORIZED FOR Property" not in operating_authority_status:
            print(f"CARRIER-VERIFICATION: Operating Authority Status verification failed: {operating_authority_status}")
            return False

        print("CARRIER-VERIFICATION: Verification successful.")
        return True
    except Exception as e:
        print(f"CARRIER-VERIFICATION: Verification failed: {e}")
        return False
    finally:
        print("========== CARRIER-VERIFICATION Section End ==========\n")



def extract_carrier_data(driver):
    print("\n========== CARRIER-DATA-EXTRACTION Section Start ==========")
    _wait = random.randint(2, 5)
    try:
        # Click on "SMS Results"
        print("CARRIER-DATA-EXTRACTION: Clicking on 'SMS Results' link...")
        sms_results_link = WebDriverWait(driver, _wait).until(
            EC.element_to_be_clickable((By.LINK_TEXT, "SMS Results"))
        )
        sms_results_link.click()

        # Scroll to "Carrier Registration Details" and click it
        print("CARRIER-DATA-EXTRACTION: Scrolling to 'Carrier Registration Details' link...")
        carrier_registration_details = WebDriverWait(driver, _wait).until(
            EC.element_to_be_clickable((By.LINK_TEXT, "Carrier Registration Details"))
        )
        ActionChains(driver).move_to_element(carrier_registration_details).perform()
        carrier_registration_details.click()

        # Switch to popup window
        print("CARRIER-DATA-EXTRACTION: Switching to popup window...")
        time.sleep(5)  # Adjust if necessary
        driver.switch_to.window(driver.window_handles[-1])

        # Extract text from the popup
        print("CARRIER-DATA-EXTRACTION: Extracting text from the popup window...")
        page_text = driver.find_element(By.TAG_NAME, "body").text

        # Regex patterns for required details
        print("CARRIER-DATA-EXTRACTION: Extracting carrier data using regex patterns...")
        legal_name_match = re.search(r"Legal Name:\s*(.+)", page_text)
        us_dot_match = re.search(r"U\.S\. DOT#:\s*(\d+)", page_text)
        #address_match = re.search(r"Address:\s*(.+?)(?=Telephone:)", page_text, re.DOTALL)
        #address_match = re.search(r"Address:\s*(.+?)\n(.+?)\nTelephone:", page_text, re.DOTALL)
        address_match = re.search(r"Address:\s*(.*?)\n(.*?)\n(?=Telephone:)", page_text, re.DOTALL)
        phone_match = re.search(r"Telephone:\s*(\(\d{3}\) \d{3}-\d{4})", page_text)
        email_match = re.search(r"Email:\s*([\w._%+-]+@[\w.-]+\.[a-zA-Z]{2,})", page_text)

        # Storing extracted data
        print("CARRIER-DATA-EXTRACTION: Storing extracted data...")
        data = {
            "Legal Name": legal_name_match.group(1).strip() if legal_name_match else "N/A",
            "U.S. DOT#": us_dot_match.group(1) if us_dot_match else "N/A",
            "Address": address_match.group(1).strip() if address_match else "N/A",
            "Telephone": phone_match.group(1) if phone_match else "N/A",
            "Email": email_match.group(1) if email_match else "N/A",
        }

        print("CARRIER-DATA-EXTRACTION: Data extraction successful.")
    except Exception as e:
        print(f"CARRIER-DATA-EXTRACTION: Error during data extraction: {e}")
        data = {
            "Legal Name": "N/A",
            "U.S. DOT#": "N/A",
            "Address": "N/A",
            "Telephone": "N/A",
            "Email": "N/A",
        }
    finally:
        print("========== CARRIER-DATA-EXTRACTION Section End ==========\n")
        return data



def send_email(driver, user_login_email, user_login_password,  recipient_email, sender_email, login_url, subject, body):
    print("\n========== EMAIL-Service Section Start ==========")
    print("EMAIL-Service: Initializing WebDriver...")
    driver.get(login_url)
    print(f"EMAIL-Service: Navigated to {login_url}")

    # Check if login page is displayed and log in
    try:
        print("EMAIL-Service: Waiting for login page to load...")
        email_input = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "rcmloginuser"))
        )
        email_input.send_keys(user_login_email)
        print("EMAIL-Service: Entered sender email.")

        password_input = driver.find_element(By.ID, "rcmloginpwd")
        password_input.send_keys(user_login_password)
        print("EMAIL-Service: Entered sender password.")

        login_button = driver.find_element(By.ID, "rcmloginsubmit")
        login_button.click()
        print("EMAIL-Service: Clicked login button.")
    except Exception as e:
        print("EMAIL-Service: Login page not displayed:", e)
        #driver.quit()
        #return False

    # Wait for login to complete
    try:
        print("EMAIL-Service: Waiting for login to complete...")
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "rcmbtn101"))
        )
        print("EMAIL-Service: Login successful.")
    except Exception as e:
        print("EMAIL-Service: Login failed:", e)
        #driver.quit()
        return False

    # Navigate directly to the compose email URL
    compose_url = "https://mail.hostinger.com/?_task=mail&_action=compose"
    driver.get(compose_url)
    print(f"EMAIL-Service: Navigated to compose email page: {compose_url}")

    # Wait for the compose window to appear
    time.sleep(2)

    # Fill in email details
    try:
        print("EMAIL-Service: Waiting for compose window to load...")
        to_input = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//input[@type='text' and @role='combobox']"))
        )
        to_input.send_keys(recipient_email)
        print("EMAIL-Service: Entered recipient email.")

        subject_input = driver.find_element(By.ID, "compose-subject")
        subject_input.send_keys(subject)
        print("EMAIL-Service: Entered email subject.")
    except Exception as e:
        print("EMAIL-Service: Failed to fill in email details:", e)
        #driver.quit()
        return False

    # Switch to the iframe for the email body
    try:
        print("EMAIL-Service: Switching to email body iframe...")
        driver.switch_to.frame(driver.find_element(By.ID, "composebody_ifr"))

        # Ensure the body contains only BMP characters
        body = ''.join(c for c in body if ord(c) <= 0xFFFF)

        # Fill in the email body
        body_input = driver.find_element(By.ID, "tinymce")
        body_input.send_keys(body)
        print("EMAIL-Service: Entered email body.")

        # Switch back to the default content
        driver.switch_to.default_content()

        time.sleep(5)  # wait a little bit to see the email body
    except Exception as e:
        print("EMAIL-Service: Failed to fill in email body:", e)
        #driver.quit()
        return False

    # Send the email
    try:
        print("EMAIL-Service: Attempting to send email...")
        send_button = driver.find_element(By.ID, "rcmbtn110")
        send_button.click()
        print("EMAIL-Service: Email sent successfully.")
        return True
    except Exception as e:
        print("EMAIL-Service: Failed to send email:", e)
        return False
    finally:
        #driver.quit()
        print("EMAIL-Service: WebDriver closed.")
        print("========== EMAIL-Service Section End ==========\n")


def save_to_excel(data, mc_mx_number, followup_status="Not Yet", file_name="carrier_data.xlsx"):
    print("\n========== STORAGE Section Start ==========")
    # Load or create workbook
    try:
        workbook = openpyxl.load_workbook(file_name)
        print("STORAGE: Workbook loaded successfully.")
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        print("STORAGE: Workbook not found. Created a new one.")

    sheet = workbook.active

    # Write headers if the sheet is empty
    if sheet.max_row == 1:
        headers = ["MC-MX Number", "Legal Name", "U.S. DOT#", "Address", "Telephone", "Email", "Followup"]
        sheet.append(headers)
        print("STORAGE: Headers written to the sheet.")

    # Write data
    try:
        sheet.append([
            mc_mx_number,
            data.get("Legal Name"),
            data.get("U.S. DOT#"),
            data.get("Address"),
            data.get("Telephone"),
            data.get("Email"),
            followup_status,
        ])
        print("STORAGE: Data appended to the sheet.")
    except Exception as e:
        print(f"STORAGE: Error appending data to the sheet: {e}")

    try:
        workbook.save(file_name)
        print(f"STORAGE: Data saved to {file_name}")
    except Exception as e:
        print(f"STORAGE: Error saving workbook: {e}")

    print("========== STORAGE Section End ==========\n")
    return data


# Main script
def main():
    # Check if MCMX_START and MCMX_END are provided
    if MCMX_START is None or MCMX_END is None:
        print("Error: MCMX_START and MCMX_END must be provided.")
        return

    # Check if email sending is enabled and email credentials are provided
    if ENABLE_EMAIL_SENDING:
        if not EMAIL_LOGIN_URL or not EMAIL_USERNAME or not EMAIL_PASSWORD:
            print("Error: EMAIL_LOGIN_URL, EMAIL_USERNAME, and EMAIL_PASSWORD must be provided when email sending is enabled.")
            return

    with open('email_data.yaml', 'r', encoding="utf-8") as file:
        email_data = yaml.safe_load(file)
    
    driver = setup_browser()
    try:
        for mc_mx_number in range(MCMX_START, MCMX_END + 1):
            print(f"Searching for MC/MX Number: {mc_mx_number}")
            if search_company(driver, mc_mx_number):
                if verify_details(driver):
                    carrier_data = extract_carrier_data(driver)
                    print('Carrier Data has been saved to excel sheet:', carrier_data)
                    
                    followup_status = "Not Yet"
                    if carrier_data.get("Email") != "N/A":
                        print(f"Sending email to {carrier_data.get('Email')}...")
                        customer_name = carrier_data.get("Legal Name").replace("LLC", "").strip()
                        email_sent = send_email(
                            driver=driver,
                            user_login_email=EMAIL_USERNAME,
                            user_login_password=EMAIL_PASSWORD,
                            recipient_email="m.shoaib_ali@outlook.com",  # carrier_data.get("Email")
                            sender_email=email_data['email_template']['sender'].strip(),
                            login_url=EMAIL_LOGIN_URL,
                            subject=email_data['email_template']['subject'].strip(),
                            body=email_data['email_template']['body'].replace('{{customer_name}}', customer_name).strip(),
                        )
                        if email_sent:
                            followup_status = "Email Sent"
                    
                    save_to_excel(carrier_data, mc_mx_number, followup_status)
                else:
                    print(f"Verification failed for MC/MX Number: {mc_mx_number}. Details do not match.")
            else: 
                print(f"Search failed for MC/MX Number: {mc_mx_number}.")
                with open("failed_to_search.txt", "a") as file:
                    file.write(f"{mc_mx_number} - Company search not found\n")

            # Optionally, add a delay between searches to avoid overwhelming the server
            time.sleep(2)

    finally:
        driver.quit()

if __name__ == "__main__":
    main()