from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
import time
import openpyxl  # For Excel integration
from selenium.webdriver.common.action_chains import ActionChains
import re
import random

# Google Sheet integration (optional)
# import gspread
# from oauth2client.service_account import ServiceAccountCredentials

def setup_browser():
    # Set up the Selenium WebDriver (Ensure you have the ChromeDriver installed and PATH set)
    driver = webdriver.Chrome()
    driver.get("https://safer.fmcsa.dot.gov/CompanySnapshot.aspx")
    return driver

def search_company(driver, number):
    _wait = random.randint(2, 5)
    # Navigate back to the search page
    driver.get("https://safer.fmcsa.dot.gov/CompanySnapshot.aspx")

    # Select the "MC/MX Number" radio button
    mc_mx_radio = WebDriverWait(driver, _wait).until(
        EC.element_to_be_clickable((By.XPATH, "//input[@id='2' and @name='query_param' and @value='MC_MX']"))
    )
    driver.execute_script("arguments[0].scrollIntoView(true);", mc_mx_radio)  # Ensure element is visible
    mc_mx_radio.click()

    # Enter the number in the search box
    search_box = WebDriverWait(driver, _wait).until(
        EC.presence_of_element_located((By.XPATH, "//input[@id='4' and @name='query_string']"))
    )
    search_box.clear()
    search_box.send_keys(number)

    # Click the "Search" button
    search_button = WebDriverWait(driver, _wait).until(
        EC.element_to_be_clickable((By.XPATH, "//input[@type='SUBMIT' and @value='Search']"))
    )
    search_button.click()


def verify_details(driver):
    _wait = random.randint(2, 5)
    # Check if "Record Not Found" is present on the page
    try:
        record_not_found = WebDriverWait(driver, _wait).until(
            EC.presence_of_element_located((By.XPATH, "//i[text()='Record Not Found']"))
        )
        print("Record Not Found. skiping further processing.")
        return False
    except Exception:
        pass

    # Check if "Record Inactive" is present on the page
    try:
        record_inactive = WebDriverWait(driver, _wait).until(
            EC.presence_of_element_located((By.XPATH, "//i[text()='Record Inactive']"))
        )
        print("Record Inactive. skipping further processing.")
        return False
    except Exception:
        pass

    # Verify the details on the loaded page
    try:
        # Verify "Entity Type"
        entity_type = WebDriverWait(driver, _wait).until(
            EC.presence_of_element_located((By.XPATH, "//th[a[text()='Entity Type:']]/following-sibling::td"))
        ).text.strip()
        if entity_type != "CARRIER":
            print(f"Entity Type verification failed: {entity_type}")
            return False

        # Verify "USDOT Status"
        usdot_status = WebDriverWait(driver, _wait).until(
            EC.presence_of_element_located((By.XPATH, "//th[a[text()='USDOT Status:']]/following-sibling::td"))
        ).text.strip()
        if usdot_status != "ACTIVE":
            print(f"USDOT Status verification failed: {usdot_status}")
            return False

        # Verify "Operating Authority Status"
        operating_authority_status = WebDriverWait(driver, _wait).until(
            EC.presence_of_element_located((By.XPATH, "//th[a[text()='Operating Authority Status:']]/following-sibling::td"))
        ).text.strip()
        if "AUTHORIZED FOR Property" not in operating_authority_status:
            print(f"Operating Authority Status verification failed: {operating_authority_status}")
            return False

        return True
    except Exception as e:
        print("Verification failed:", e)
        return False

def extract_carrier_data(driver):
    _wait= random.randint(2, 5)
    # Click on "SMS Results"
    sms_results_link = WebDriverWait(driver, _wait).until(
        EC.element_to_be_clickable((By.LINK_TEXT, "SMS Results"))
    )
    sms_results_link.click()

    # Scroll to "Carrier Registration Details" and click it
    carrier_registration_details = WebDriverWait(driver, _wait).until(
        EC.element_to_be_clickable((By.LINK_TEXT, "Carrier Registration Details"))
    )
    ActionChains(driver).move_to_element(carrier_registration_details).perform()
    carrier_registration_details.click()

    # Switch to popup window
    time.sleep(5)  # Adjust if necessary
    driver.switch_to.window(driver.window_handles[-1])

    # Extract text from the popup
    page_text = driver.find_element(By.TAG_NAME, "body").text

    # Regex patterns for required details
    legal_name_match = re.search(r"Legal Name:\s*(.+)", page_text)
    us_dot_match = re.search(r"U\.S\. DOT#:\s*(\d+)", page_text)
    #address_match = re.search(r"Address:\s*(.+?)(?=Telephone:)", page_text, re.DOTALL)
    #address_match = re.search(r"Address:\s*(.+?)\n(.+?)\nTelephone:", page_text, re.DOTALL)
    address_match = re.search(r"Address:\s*(.*?)\n(.*?)\n(?=Telephone:)", page_text, re.DOTALL)
    phone_match = re.search(r"Telephone:\s*(\(\d{3}\) \d{3}-\d{4})", page_text)
    email_match = re.search(r"Email:\s*([\w._%+-]+@[\w.-]+\.[a-zA-Z]{2,})", page_text)

    # Storing extracted data
    data = {
        "Legal Name": legal_name_match.group(1).strip() if legal_name_match else "N/A",
        "U.S. DOT#": us_dot_match.group(1) if us_dot_match else "N/A",
        "Address": address_match.group(1).strip() if address_match else "N/A",
        "Telephone": phone_match.group(1) if phone_match else "N/A",
        "Email": email_match.group(1) if email_match else "N/A",
    }
 
    return data

def save_to_excel(data, mc_mx_number, file_name="carrier_data.xlsx"):
    # Load or create workbook
    try:
        workbook = openpyxl.load_workbook(file_name)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    sheet = workbook.active

    # Write headers if the sheet is empty
    if sheet.max_row == 1:
        headers = ["MC-MX Number", "Legal Name", "U.S. DOT#", "Address", "Telephone", "Email", "Followup"]
        sheet.append(headers)

    # Write data
    sheet.append([
        mc_mx_number,
        data.get("Legal Name"),
        data.get("U.S. DOT#"),
        data.get("Address"),
        data.get("Telephone"),
        data.get("Email"),
        "Not Yet",  
    ])

    workbook.save(file_name)
    print(f"Data saved to {file_name}")


# Main script
def main():
    driver = setup_browser()

    try:
        # Define the range of numbers to search
        start_number = 1635000
        end_number = 16435020
  
        for mc_mx_number in range(start_number, end_number + 1):
            #time.sleep(random.randint(2, 5))
            print(f"Searching for MC/MX Number: {mc_mx_number}")
            search_company(driver, mc_mx_number)

            if verify_details(driver):
                carrier_data = extract_carrier_data(driver)
                print('Carrier Data has been saved to excel sheet:', carrier_data)
                save_to_excel(carrier_data, mc_mx_number)
            else:
                print(f"Verification failed for MC/MX Number: {mc_mx_number}. Details do not match.")

            # Optionally, add a delay between searches to avoid overwhelming the server
            time.sleep(2)

    finally:
        driver.quit()

if __name__ == "__main__":
    main()
